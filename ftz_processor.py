"""
Unimex Customs - FTZ Separation Processor
=========================================
For each shipment (identified by a carrier bill-of-lading number, e.g.
COSU..., EGLV...), pairs the master manifest with its FTZ separation list,
filters the master to only line items whose Bag ID appears in the separation
list, then aggregates to one row per HS Code with summed Quantity, Weight,
and Value.

Run:
    python ftz_processor.py                 # reads ./input, writes ./output
    python ftz_processor.py /path/in /path/out

Drop both files (master + separation list) for any number of shipments into the
input folder. The script pairs them up by the shipment ID (SCAC + serial bill
of lading) found in each file's contents -- filenames are not trusted.
"""

from __future__ import annotations

import os
import re
import sys
import traceback
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MAX_ROWS_PER_SHEET = 998

# Master manifest columns -- exact names as they appear in the client's export.
# Each master column is matched against a list of accepted aliases (primary),
# then against keyword tokens as a fallback (secondary). Aliases are compared
# case-insensitively after stripping all whitespace. Add new aliases here when
# clients send files with slightly different headers.
COL_ALIASES: dict[str, list[str]] = {
    "MWB":          ["MWB"],
    "Bag ID":       ["Bag ID"],
    "Tracking #":   ["Tracking Number"],
    "HS Code":      ["HS CODE", "HSCODE"],
    "Weight":       ["PARCEL WEIGHT", "WEIGHT"],
    "Quantity":     ["PRODUCT QTY", "TOTAL QTY", "QTY"],
    "Value":        ["TOTAL DECLARE VALUE"],
}

# Fallback keyword tokens: if no alias matches exactly, any header that
# contains ALL tokens for a column (after normalizing) is accepted.
# Tokens should be specific enough to avoid false positives.
COL_KEYWORDS: dict[str, list[str]] = {
    "MWB":        ["mwb"],
    "Bag ID":     ["bag"],
    "Tracking #": ["track"],
    "HS Code":    ["hs"],
    "Weight":     ["weight"],
    "Quantity":   ["qty"],
    "Value":      ["value"],
}

# Master sheets to try, in order of preference. '表1' is the full export;
# '0' is the same information in a different format.
MASTER_SHEET_PREFERENCE = ["表1", "0"]

# ID patterns -- used to identify file role by content, not filename.
RE_BAG      = re.compile(r"^ZXWR\d+$")
RE_TRACKING = re.compile(r"^JMX\d+$")
RE_SHIPMENT_ID = re.compile(r"\b[A-Z]{4}\d{8,}\b")


def _norm_header(h: object) -> str:
    """Lowercase + strip all whitespace, so 'HS CODE' == 'hscode' == 'hs code'."""
    if h is None:
        return ""
    return "".join(str(h).lower().split())


def resolve_columns(df_columns: list[str]) -> dict[str, str] | None:
    """
    Map each canonical column (MWB, Bag ID, etc.) to the actual column name in
    the file. Returns None if any required column can't be found.

    Matching runs in two passes:
      1. Exact alias match (case-insensitive, whitespace-stripped).
      2. Keyword fallback: any header whose normalized form contains all
         tokens for that column (e.g. any header with "weight" maps to Weight).
         The shortest matching header wins to avoid grabbing a catch-all column.
    """
    norm_to_actual: dict[str, str] = {_norm_header(c): str(c).strip() for c in df_columns}
    resolved: dict[str, str] = {}
    for canonical, aliases in COL_ALIASES.items():
        match = None

        # Pass 1: exact alias
        for alias in aliases:
            actual = norm_to_actual.get(_norm_header(alias))
            if actual:
                match = actual
                break

        # Pass 2: keyword fallback
        if not match:
            tokens = COL_KEYWORDS.get(canonical, [])
            candidates = [
                actual for norm, actual in norm_to_actual.items()
                if all(t in norm for t in tokens)
            ]
            if candidates:
                # prefer the shortest header (least likely to be a catch-all)
                match = min(candidates, key=len)
                print(f"    [INFO] '{canonical}' matched via keyword fallback -> '{match}'")

        if not match:
            return None  # missing required column
        resolved[canonical] = match
    return resolved


# ---------------------------------------------------------------------------
# Normalization
# ---------------------------------------------------------------------------
def clean_id(value: object) -> str:
    """
    Normalize an ID for matching: strip every kind of whitespace including
    \xa0 non-breaking space. Returns '' for NaN/None.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = unicodedata.normalize("NFKC", str(value))
    return "".join(s.split())


# ---------------------------------------------------------------------------
# File classification
# ---------------------------------------------------------------------------
@dataclass
class MasterFile:
    path: Path
    shipment_id: str
    sheet_name: str

@dataclass
class SeparationFile:
    path: Path
    shipment_id: str
    bag_ids: set[str]
    tracking_numbers: set[str]  # diagnostic only; not used for filtering


def classify_file(path: Path) -> MasterFile | SeparationFile | None:
    """Decide whether path is a master manifest or a separation list."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [WARN] cannot open {path.name}: {e}")
        return None

    # --- Master? look for the canonical header in a preferred sheet ---
    for sheet_name in MASTER_SHEET_PREFERENCE:
        if sheet_name not in wb.sheetnames:
            continue
        try:
            df_head = pd.read_excel(path, sheet_name=sheet_name, nrows=2, dtype=object)
        except Exception:
            continue
        resolved = resolve_columns(list(df_head.columns))
        if resolved is not None:
            shipment_id = ""
            mwb_col = resolved["MWB"]
            if len(df_head) > 0 and mwb_col in df_head.columns:
                shipment_id = clean_id(df_head.iloc[0][mwb_col])
            if not shipment_id:
                m = RE_SHIPMENT_ID.search(path.name)
                shipment_id = m.group(0) if m else ""
            if shipment_id:
                return MasterFile(path=path, shipment_id=shipment_id, sheet_name=sheet_name)

    # --- Separation list? scan cells for ZXWR/JMX patterns ---
    bags: set[str] = set()
    tracks: set[str] = set()
    shipment_id_found = ""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                s = clean_id(cell)
                if RE_BAG.match(s):
                    bags.add(s)
                elif RE_TRACKING.match(s):
                    tracks.add(s)
                elif not shipment_id_found:
                    m = RE_SHIPMENT_ID.search(s)
                    if m:
                        shipment_id_found = m.group(0)

    if bags:  # contains at least one ZXWR... = separation list
        if not shipment_id_found:
            m = RE_SHIPMENT_ID.search(path.name)
            shipment_id_found = m.group(0) if m else ""
        return SeparationFile(
            path=path,
            shipment_id=shipment_id_found,
            bag_ids=bags,
            tracking_numbers=tracks,
        )

    return None


# ---------------------------------------------------------------------------
# Processing
# ---------------------------------------------------------------------------
def process_shipment(master: MasterFile, sep: SeparationFile, out_path: Path) -> str:
    """Filter master by Bag ID, aggregate to one row per HS Code, write output."""
    df = pd.read_excel(master.path, sheet_name=master.sheet_name, dtype=object)
    df.columns = [str(c).strip() for c in df.columns]
    resolved = resolve_columns(list(df.columns))
    if resolved is None:
        raise ValueError("master file is missing one of the required columns "
                         "(MWB, Bag ID, Tracking Number, HS Code, Weight, "
                         "Quantity, Value) under any known alias")

    # Normalize the bag column for matching
    df["_bag_clean"] = df[resolved["Bag ID"]].map(clean_id)
    sep_bags_clean = {clean_id(b) for b in sep.bag_ids}

    # Sanity: any separation bags missing from the master?
    master_bags = set(df["_bag_clean"]) - {""}
    missing = sep_bags_clean - master_bags
    if missing:
        print(f"    [WARN] {len(missing)} bag ID(s) in separation list "
              f"not found in master: {sorted(missing)[:5]}"
              f"{' ...' if len(missing) > 5 else ''}")

    # Filter to FTZ rows by Bag ID
    hits = df[df["_bag_clean"].isin(sep_bags_clean)].copy()
    if hits.empty:
        return "no matching rows -- nothing to write"

    # Aggregate one row per HS Code -- sum qty, weight, value across line items.
    # Customs requires that any FINAL aggregated value < $1 or weight < 1 kg
    # be rounded up to 1. This is applied AFTER summing, not per line item,
    # so an HS code with three $0.40 line items totals $1.20 (stays as-is),
    # not $3.00.
    hits["_hs"]  = hits[resolved["HS Code"]].map(clean_id)
    hits["_qty"] = pd.to_numeric(hits[resolved["Quantity"]], errors="coerce")
    hits["_wt"]  = pd.to_numeric(hits[resolved["Weight"]],   errors="coerce")
    hits["_val"] = pd.to_numeric(hits[resolved["Value"]],    errors="coerce")
    hits = hits[hits["_hs"] != ""]

    agg = (
        hits.groupby("_hs", as_index=False)
            .agg(Quantity=("_qty", "sum"),
                 Weight  =("_wt",  "sum"),
                 Value   =("_val", "sum"))
            .rename(columns={"_hs": "HS Code"})
            .sort_values("HS Code", kind="stable")
            .reset_index(drop=True)
    )

    # Apply the round-up-to-1 rule to the aggregated Weight and Value.
    agg["Weight"] = agg["Weight"].clip(lower=1)
    agg["Value"]  = agg["Value"].clip(lower=1)

    # Constant columns required by customs: Zone = 'P', Charges = 3,
    # Country of Origin = 'CN' (always China).
    agg["Zone"]    = "P"
    agg["Charges"] = 3
    agg["Country of Origin"] = "CN"

    # Final column order
    agg = agg[["HS Code", "Quantity", "Weight", "Value",
               "Zone", "Charges", "Country of Origin"]]

    # Split into <= 998-row sheets. With aggregation this is almost never
    # triggered, but it's part of the spec.
    if len(agg) <= MAX_ROWS_PER_SHEET:
        chunks = [agg]
    else:
        chunks = [agg.iloc[i:i + MAX_ROWS_PER_SHEET].reset_index(drop=True)
                  for i in range(0, len(agg), MAX_ROWS_PER_SHEET)]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for i, chunk in enumerate(chunks, start=1):
            sheet_name = f"FTZ_{i}" if len(chunks) > 1 else "FTZ"
            chunk.to_excel(writer, sheet_name=sheet_name, index=False)

    format_workbook(out_path)

    total_value  = agg["Value"].sum()
    total_weight = agg["Weight"].sum()
    sheets_msg = f", {len(chunks)} sheets" if len(chunks) > 1 else ""
    return (f"{len(hits)} line items -> {len(agg)} HS codes"
            f"{sheets_msg}, total ${total_value:,.2f} / {total_weight:,.2f} kg")


# ---------------------------------------------------------------------------
# Output formatting
# ---------------------------------------------------------------------------
def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    fmt_by_name = {
        "Quantity": "#,##0",
        "Weight":   "#,##0.000",
        "Value":    "$#,##0.00",
        "Charges":  "#,##0",
    }
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        headers = [c.value for c in ws[1]]
        for idx, name in enumerate(headers, start=1):
            fmt = fmt_by_name.get(name)
            if fmt:
                for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                    for cell in row:
                        cell.number_format = fmt
        for idx in range(1, ws.max_column + 1):
            letter = get_column_letter(idx)
            max_len = max(
                (len(str(c.value)) for c in ws[letter] if c.value is not None),
                default=10,
            )
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 40)
        ws.freeze_panes = "A2"
    wb.save(path)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def base_dir() -> Path:
    """
    Folder the script (or .exe) lives in. When packaged with PyInstaller and
    double-clicked, the current working directory can be C:\\Windows\\System32,
    so we always anchor input/output to the .exe's actual location.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


class TeeLogger:
    """Write everything printed to console to a log file as well."""
    def __init__(self, log_path: Path):
        self.terminal = sys.stdout
        self.log = open(log_path, "w", encoding="utf-8")
    def write(self, msg: str) -> None:
        self.terminal.write(msg)
        self.log.write(msg)
        self.log.flush()
    def flush(self) -> None:
        self.terminal.flush()
        self.log.flush()
    def close(self) -> None:
        try: self.log.close()
        except Exception: pass


def pause_for_user() -> None:
    """
    Wait for a keypress so the console window doesn't flash shut when run by
    double-click. No-op when redirected (e.g. running in CI).
    """
    if not sys.stdout.isatty():
        return
    try:
        input("\nPress ENTER to close...")
    except (EOFError, KeyboardInterrupt):
        pass


def run(in_dir: Path, out_dir: Path) -> int:
    """The actual processing -- separated so main() can wrap it in try/except."""
    if not in_dir.exists():
        print(f"Creating input folder: {in_dir}")
        in_dir.mkdir(parents=True, exist_ok=True)
        print(f"\nNo input files yet. Drop the master and separation files into:")
        print(f"  {in_dir}")
        print(f"...then run this again.")
        return 0
    out_dir.mkdir(parents=True, exist_ok=True)

    files = sorted(
        p for p in in_dir.iterdir()
        if p.suffix.lower() in {".xlsx", ".xls", ".xlsm"} and not p.name.startswith("~$")
    )
    if not files:
        print(f"No Excel files found in: {in_dir}")
        print(f"Drop your shipment files there and run again.")
        return 0

    print(f"Scanning {len(files)} file(s) in: {in_dir}\n")

    masters: dict[str, MasterFile] = {}
    separations: dict[str, SeparationFile] = {}
    unrecognized: list[Path] = []

    for path in files:
        result = classify_file(path)
        if isinstance(result, MasterFile):
            print(f"  [MASTER] {path.name}  ->  {result.shipment_id}")
            if result.shipment_id in masters:
                print(f"    [WARN] duplicate master for {result.shipment_id}; overwriting")
            masters[result.shipment_id] = result
        elif isinstance(result, SeparationFile):
            print(f"  [SEP   ] {path.name}  ->  {result.shipment_id}  "
                  f"({len(result.bag_ids)} bag(s))")
            if result.shipment_id in separations:
                print(f"    [WARN] duplicate separation list for {result.shipment_id}; overwriting")
            separations[result.shipment_id] = result
        else:
            print(f"  [?????] {path.name}  (not recognized -- check the file)")
            unrecognized.append(path)

    print()
    all_shipment_ids = set(masters) | set(separations)
    processed = skipped = 0
    for shipment_id in sorted(all_shipment_ids):
        if shipment_id not in masters:
            print(f"  [SKIP] {shipment_id}: separation list found but no master manifest")
            skipped += 1
            continue
        if shipment_id not in separations:
            print(f"  [SKIP] {shipment_id}: master found but no separation list")
            skipped += 1
            continue
        out_path = out_dir / f"{shipment_id}_FTZ.xlsx"
        try:
            msg = process_shipment(masters[shipment_id], separations[shipment_id], out_path)
            print(f"  [OK  ] {shipment_id} -> {out_path.name}")
            print(f"         {msg}")
            processed += 1
        except Exception as e:
            print(f"  [FAIL] {shipment_id}: {e}")
            skipped += 1

    print(f"\n{'=' * 60}")
    print(f"Done. {processed} shipment(s) processed, {skipped} skipped.")
    if unrecognized:
        print(f"{len(unrecognized)} file(s) not recognized.")
    print(f"Output folder: {out_dir}")
    print('=' * 60)
    return 0


def main() -> int:
    here = base_dir()
    in_dir  = Path(sys.argv[1]) if len(sys.argv) > 1 else here / "input"
    out_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else here / "output"

    # Log file lives next to the .exe with a timestamp
    log_dir = here / "logs"
    log_dir.mkdir(exist_ok=True)
    log_path = log_dir / f"run_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    tee = TeeLogger(log_path)
    sys.stdout = tee
    sys.stderr = tee

    print(f"Unimex FTZ Processor")
    print(f"Run started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Log file: {log_path}")
    print()

    exit_code = 1
    try:
        exit_code = run(in_dir, out_dir)
    except Exception as e:
        print(f"\n[ERROR] Something went wrong: {e}")
        print(f"\nFull details (please send this log file to Andy):")
        traceback.print_exc()
    finally:
        sys.stdout = tee.terminal
        sys.stderr = tee.terminal
        tee.close()
        pause_for_user()
    return exit_code


if __name__ == "__main__":
    raise SystemExit(main())
