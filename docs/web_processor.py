"""
Browser-side wrapper for the FTZ processor. Runs inside Pyodide.

Re-uses the helpers and constants from ftz_processor.py (the canonical CLI
script). The functions here mirror classify_file / process_shipment but
operate on in-memory bytes instead of disk paths.

JS entry point: process_files(files) where files is a list of dicts with
keys "name" (str) and "data" (bytes / Uint8Array).
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ftz_processor import (
    MAX_ROWS_PER_SHEET,
    MASTER_SHEET_PREFERENCE,
    RE_BAG,
    RE_TRACKING,
    RE_SHIPMENT_ID,
    clean_id,
    resolve_columns,
)


@dataclass
class _Master:
    name: str
    data: bytes
    shipment_id: str
    sheet_name: str


@dataclass
class _Sep:
    name: str
    data: bytes
    shipment_id: str
    bag_ids: set
    tracking_numbers: set


def _classify(name: str, data: bytes):
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        return None, f"cannot open: {e}"

    for sheet_name in MASTER_SHEET_PREFERENCE:
        if sheet_name not in wb.sheetnames:
            continue
        try:
            df_head = pd.read_excel(
                io.BytesIO(data), sheet_name=sheet_name, nrows=2, dtype=object
            )
        except Exception:
            continue
        resolved = resolve_columns(list(df_head.columns))
        if resolved is not None:
            shipment_id = ""
            mwb_col = resolved["MWB"]
            if len(df_head) > 0 and mwb_col in df_head.columns:
                shipment_id = clean_id(df_head.iloc[0][mwb_col])
            if not shipment_id:
                m = RE_SHIPMENT_ID.search(name)
                shipment_id = m.group(0) if m else ""
            if shipment_id:
                return _Master(name=name, data=data,
                               shipment_id=shipment_id,
                               sheet_name=sheet_name), None

    bags: set = set()
    tracks: set = set()
    shipment_id_found = ""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                s = clean_id(cell)
                if RE_BAG.match(s):
                    bags.add(s)
                elif RE_TRACKING.match(s):
                    tracks.add(s)
                elif not shipment_id_found:
                    m = RE_SHIPMENT_ID.search(s)
                    if m:
                        shipment_id_found = m.group(0)

    if bags:
        if not shipment_id_found:
            m = RE_SHIPMENT_ID.search(name)
            shipment_id_found = m.group(0) if m else ""
        return _Sep(name=name, data=data,
                    shipment_id=shipment_id_found,
                    bag_ids=bags, tracking_numbers=tracks), None

    return None, "not recognized as master or separation list"


def _format_wb(wb) -> None:
    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    fmt_by_name = {
        "Quantity": "#,##0",
        "Weight":   "#,##0.000",
        "Value":    "$#,##0.00",
        "Charges":  "#,##0",
    }
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        headers = [c.value for c in ws[1]]
        for idx, name in enumerate(headers, start=1):
            fmt = fmt_by_name.get(name)
            if fmt:
                for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                    for cell in row:
                        cell.number_format = fmt
        for idx in range(1, ws.max_column + 1):
            letter = get_column_letter(idx)
            max_len = max(
                (len(str(c.value)) for c in ws[letter] if c.value is not None),
                default=10,
            )
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 40)
        ws.freeze_panes = "A2"


def _process(master: _Master, sep: _Sep):
    warnings: list[str] = []
    df = pd.read_excel(io.BytesIO(master.data),
                       sheet_name=master.sheet_name, dtype=object)
    df.columns = [str(c).strip() for c in df.columns]
    resolved = resolve_columns(list(df.columns))
    if resolved is None:
        raise ValueError("master file is missing one of the required columns "
                         "(MWB, Bag ID, Tracking Number, HS Code, Weight, "
                         "Quantity, Value) under any known alias")

    df["_bag_clean"] = df[resolved["Bag ID"]].map(clean_id)
    sep_bags_clean = {clean_id(b) for b in sep.bag_ids}

    master_bags = set(df["_bag_clean"]) - {""}
    missing = sep_bags_clean - master_bags
    if missing:
        sample = sorted(missing)[:5]
        suffix = " ..." if len(missing) > 5 else ""
        warnings.append(f"{len(missing)} bag ID(s) in separation list "
                        f"not found in master: {sample}{suffix}")

    hits = df[df["_bag_clean"].isin(sep_bags_clean)].copy()
    if hits.empty:
        return None, "no matching rows -- nothing to write", warnings

    hits["_hs"]  = hits[resolved["HS Code"]].map(clean_id)
    hits["_qty"] = pd.to_numeric(hits[resolved["Quantity"]], errors="coerce")
    hits["_wt"]  = pd.to_numeric(hits[resolved["Weight"]],   errors="coerce")
    hits["_val"] = pd.to_numeric(hits[resolved["Value"]],    errors="coerce")
    hits = hits[hits["_hs"] != ""]

    agg = (
        hits.groupby("_hs", as_index=False)
            .agg(Quantity=("_qty", "sum"),
                 Weight  =("_wt",  "sum"),
                 Value   =("_val", "sum"))
            .rename(columns={"_hs": "HS Code"})
            .sort_values("HS Code", kind="stable")
            .reset_index(drop=True)
    )
    agg["Weight"] = agg["Weight"].clip(lower=1)
    agg["Value"]  = agg["Value"].clip(lower=1)
    agg["Zone"]    = "P"
    agg["Charges"] = 3
    agg["Country of Origin"] = "CN"
    agg = agg[["HS Code", "Quantity", "Weight", "Value",
               "Zone", "Charges", "Country of Origin"]]

    if len(agg) <= MAX_ROWS_PER_SHEET:
        chunks = [agg]
    else:
        chunks = [agg.iloc[i:i + MAX_ROWS_PER_SHEET].reset_index(drop=True)
                  for i in range(0, len(agg), MAX_ROWS_PER_SHEET)]

    raw_buf = io.BytesIO()
    with pd.ExcelWriter(raw_buf, engine="openpyxl") as writer:
        for i, chunk in enumerate(chunks, start=1):
            sheet_name = f"FTZ_{i}" if len(chunks) > 1 else "FTZ"
            chunk.to_excel(writer, sheet_name=sheet_name, index=False)

    raw_buf.seek(0)
    wb = load_workbook(raw_buf)
    _format_wb(wb)
    out_buf = io.BytesIO()
    wb.save(out_buf)

    total_value  = float(agg["Value"].sum())
    total_weight = float(agg["Weight"].sum())
    sheets_msg = f", {len(chunks)} sheets" if len(chunks) > 1 else ""
    summary = (f"{len(hits)} line items → {len(agg)} HS codes"
               f"{sheets_msg}, total ${total_value:,.2f} / "
               f"{total_weight:,.2f} kg")

    preview_rows = [
        [_jsonify(v) for v in row]
        for row in agg.itertuples(index=False, name=None)
    ]
    return {
        "xlsx": out_buf.getvalue(),
        "preview_headers": list(agg.columns),
        "preview_rows": preview_rows,
        "summary": summary,
    }, summary, warnings


def _jsonify(v):
    if v is None:
        return ""
    if isinstance(v, float):
        if pd.isna(v):
            return ""
        return v
    return v


def process_files(files):
    """
    files: iterable of {"name": str, "data": bytes-like}
    Returns a dict: {results, skipped, unrecognized}
    """
    masters: dict[str, _Master] = {}
    seps: dict[str, _Sep] = {}
    unrecognized: list[dict] = []

    # Normalize Pyodide JsProxy / Python dicts to plain values.
    for f in files:
        if hasattr(f, "to_py"):
            f = f.to_py()
        name = f["name"]
        raw = f["data"]
        if not isinstance(raw, (bytes, bytearray)):
            raw = bytes(raw)

        result, err = _classify(name, raw)
        if isinstance(result, _Master):
            if result.shipment_id in masters:
                # duplicate master -- keep newest, surface warning later
                pass
            masters[result.shipment_id] = result
        elif isinstance(result, _Sep):
            seps[result.shipment_id] = result
        else:
            unrecognized.append({"name": name, "reason": err or "unknown"})

    results = []
    skipped = []
    for sid in sorted(set(masters) | set(seps)):
        if sid not in masters:
            skipped.append({"shipment_id": sid,
                            "reason": "separation list found but no master"})
            continue
        if sid not in seps:
            skipped.append({"shipment_id": sid,
                            "reason": "master found but no separation list"})
            continue
        try:
            out, summary, warnings = _process(masters[sid], seps[sid])
            if out is None:
                skipped.append({"shipment_id": sid, "reason": summary})
                continue
            results.append({
                "shipment_id": sid,
                "filename": f"{sid}_FTZ.xlsx",
                "xlsx": out["xlsx"],
                "preview_headers": out["preview_headers"],
                "preview_rows": out["preview_rows"],
                "summary": summary,
                "warnings": warnings,
            })
        except Exception as e:
            skipped.append({"shipment_id": sid, "reason": str(e)})

    return {
        "results": results,
        "skipped": skipped,
        "unrecognized": unrecognized,
    }
